import streamlit as st
import base64
import time
import hmac
import hashlib
import urllib.parse
import requests
import pandas as pd
import json
import io
import pytz
import datetime as dt
from datetime import datetime, timedelta, time as dt_time, date
from supabase import create_client, Client

# ===============================
# TIMEZONE SETUP - TAMBAHKAN SETELAH IMPORT
# ===============================
WIB = pytz.timezone('Asia/Jakarta')
UTC = pytz.UTC

import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Di awal setiap tab, log timezone info
logger.info(f"Server time (UTC): {datetime.now(pytz.UTC)}")
logger.info(f"WIB time: {datetime.now(WIB)}")


# ===============================
# PAGE CONFIG
# ===============================
st.set_page_config(page_title="Humanoid Shopee API", layout="wide")

# ===============================
# SUPABASE CONFIG
# ===============================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ===============================
# SHOPEE CONFIG
# ===============================
PARTNER_ID = st.secrets.get("PARTNER_ID", "")
PARTNER_KEY = st.secrets.get("PARTNER_KEY", "")
REDIRECT_URL = st.secrets.get("REDIRECT_URL", "")

BASE_URL = "https://partner.shopeemobile.com"

# ===============================
# CAPTURE REDIRECT PARAMS (OAUTH)
# ===============================
query_params = st.query_params

oauth_code = query_params.get("code")
oauth_shop_id = query_params.get("shop_id")

# ===============================
# SIGNATURE HELPERS
# ===============================
def generate_sign_basic(path, timestamp):
    base_string = f"{PARTNER_ID}{path}{timestamp}"
    return hmac.new(
        PARTNER_KEY.encode(),
        base_string.encode(),
        hashlib.sha256
    ).hexdigest()

def generate_sign_full(path, timestamp, access_token, shop_id):
    base_string = f"{PARTNER_ID}{path}{timestamp}{access_token}{shop_id}"
    return hmac.new(
        PARTNER_KEY.encode(),
        base_string.encode(),
        hashlib.sha256
    ).hexdigest()

# ===============================
# DB HELPERS
# ===============================
def save_token_to_db(shop_name, shop_id, access_token, refresh_token):
    data = {
        "shop_name": shop_name,
        "shop_id": int(shop_id),
        "access_token": access_token,
        "refresh_token": refresh_token,
        "updated_at": "now()"
    }
    supabase.table("shopee_tokens").upsert(data).execute()

def get_all_shops():
    res = supabase.table("shopee_tokens").select("shop_name").execute()
    return [r["shop_name"] for r in res.data] if res.data else []

def get_shop_token(shop_name):
    res = supabase.table("shopee_tokens").select("*").eq("shop_name", shop_name).execute()
    if not res.data:
        return None
    return res.data[0]

def save_report_to_db(shop_name, date_range, excel_bytes):
    excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
    data = {
        "shop_name": shop_name,
        "date_range": date_range,
        "csv_content": excel_base64,
        "created_at": "now()"
    }
    try:
        supabase.table("shopee_reports").insert(data).execute()
    except Exception as e:
        st.error(f"Gagal simpan ke Database: {e}")
        raise e

def get_report_history(shop_name):
    res = supabase.table("shopee_reports").select("*").eq("shop_name", shop_name).order("created_at", desc=True).limit(10).execute()
    return res.data
    
# ===============================
# TOKEN REFRESH
# ===============================
def auto_refresh_token(shop_name):
    token_data = get_shop_token(shop_name)
    if not token_data:
        return None, None
    
    path = "/api/v2/auth/access_token/get"
    timestamp = int(time.time())
    sign = generate_sign_basic(path, timestamp)
    
    url = BASE_URL + path
    payload = {
        "partner_id": int(PARTNER_ID),
        "refresh_token": token_data['refresh_token'],
        "shop_id": int(token_data['shop_id']),
        "timestamp": timestamp,
        "sign": sign
    }
    
    r = requests.post(url, json=payload).json()
    
    if "access_token" in r.get("response", {}):
        new_at = r["response"]["access_token"]
        new_rt = r["response"]["refresh_token"]
        
        save_token_to_db(shop_name, token_data['shop_id'], new_at, new_rt)
        return new_at, token_data['shop_id']
    else:
        st.error(f"Gagal Refresh Token: {r}")
        return None, None

def get_escrow_detail(order_sn, access_token, shop_id):
    path = "/api/v2/payment/get_escrow_detail"
    timestamp = int(time.time())
    sign = generate_sign_full(path, timestamp, access_token, shop_id)
    
    url = BASE_URL + path
    params = {
        "partner_id": PARTNER_ID,
        "timestamp": timestamp,
        "access_token": access_token,
        "shop_id": int(shop_id),
        "sign": sign,
        "order_sn": order_sn
    }
    r = requests.get(url, params=params).json()
    return r.get("response", {})

def create_income_excel(df_inc, df_srv, df_prc, shop_name, start_date, end_date):
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # --- SHEET 1: Summary dengan Format Rapi ---
        ws_summary = workbook.create_sheet('Summary')
        
        # Style untuk formatting
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # Bold font untuk header
        bold_font = Font(bold=True, size=11)
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=11)
        
        # Alignment
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Border
        thin_border = Border(
            bottom=Side(style='thin', color='000000')
        )
        
        # Helper untuk menulis row
        def write_row(row_idx, col1_text, col2_text="", col3_text="", bold=False, indent=0):
            cell1 = ws_summary.cell(row=row_idx, column=1, value=col1_text)
            cell2 = ws_summary.cell(row=row_idx, column=2, value=col2_text)
            cell3 = ws_summary.cell(row=row_idx, column=3, value=col3_text)
            
            if bold:
                cell1.font = bold_font
                cell2.font = bold_font
                cell3.font = bold_font
            
            cell1.alignment = Alignment(horizontal='left', vertical='center', indent=indent)
            cell2.alignment = right_align
            cell3.alignment = right_align
            
            return row_idx + 1

        # Hitung nilai-nilai
        harga_asli = df_inc["Harga Asli Produk"].sum()
        total_diskon = df_inc["Total Diskon Produk"].sum()
        refund_pembeli = df_inc["Jumlah Pengembalian Dana ke Pembeli"].sum()
        subtotal_pesanan = harga_asli + total_diskon - refund_pembeli
        
        # Voucher & Subsidi
        diskon_shopee = df_inc["Diskon Produk dari Shopee"].sum()
        voucher_penjual = df_inc["Voucher dari Penjual"].sum()
        cashback_penjual = df_inc["Cashback Koin dari Penjual"].sum()
        total_voucher_subsidi = diskon_shopee + voucher_penjual + cashback_penjual
        
        # Biaya Pengiriman
        ongkir_pembeli = df_inc["Ongkir Dibayar Pembeli"].sum()
        diskon_ongkir_3pl = df_inc["Diskon Ongkir Ditanggung Jasa Kirim"].sum()
        gratis_ongkir_shopee = df_inc["Gratis Ongkir dari Shopee"].sum()
        ongkir_diteruskan = df_inc["Ongkir yang Diteruskan oleh Shopee ke Jasa Kirim"].sum()
        ongkir_return = df_inc["Ongkos Kirim Pengembalian Barang"].sum()
        pengembalian_biaya_kirim = df_inc["Pengembalian Biaya Kirim"].sum()
        kembali_biaya_pengirim = df_inc["Kembali ke Biaya Pengiriman Pengirim"].sum()
        total_biaya_pengiriman = (ongkir_pembeli + diskon_ongkir_3pl + gratis_ongkir_shopee + 
                                  ongkir_diteruskan + ongkir_return + pengembalian_biaya_kirim + 
                                  kembali_biaya_pengirim)
        
        # Biaya Admin & Layanan
        biaya_ams = df_inc["Biaya Komisi AMS"].sum()
        biaya_admin = df_inc["Biaya Administrasi"].sum()
        biaya_layanan = df_inc["Biaya Layanan"].sum()
        biaya_proses = df_inc["Biaya Proses Pesanan"].sum()
        premi = df_inc["Premi"].sum()
        biaya_hemat_kirim = df_inc["Biaya Program Hemat Biaya Kirim"].sum()
        biaya_transaksi = df_inc["Biaya Transaksi"].sum()
        biaya_kampanye = df_inc["Biaya Kampanye"].sum()
        total_biaya_admin = biaya_ams + biaya_admin + biaya_layanan + biaya_proses + premi + biaya_hemat_kirim + biaya_transaksi + biaya_kampanye
        
        # Total perhitungan
        total_pendapatan = subtotal_pesanan + total_voucher_subsidi
        total_pengeluaran = total_biaya_pengiriman + total_biaya_admin
        total_dilepas = df_inc["Total Penghasilan"].sum()
        
        # Promo dari penjual
        promo_gratis_ongkir = df_inc["Promo Gratis Ongkir dari Penjual"].sum()
        
        # Mulai menulis data
        current_row = 1
        
        # Header
        current_row = write_row(current_row, "Laporan Penghasilan", "", "", bold=True)
        current_row += 1  # Empty row
        
        # Rincian Laporan
        current_row = write_row(current_row, "Rincian Laporan", "", "", bold=True)
        current_row = write_row(current_row, "Username (Penjual)", shop_name)
        current_row = write_row(current_row, "Dari", str(start_date))
        current_row = write_row(current_row, "ke", str(end_date))
        current_row += 1  # Empty row
        
        # Ringkasan Penghasilan
        current_row = write_row(current_row, "Ringkasan Penghasilan", "", "Rp", bold=True)
        
        # 1. Total Pendapatan
        current_row = write_row(current_row, "1. Total Pendapatan", "", total_pendapatan, bold=True)
        current_row = write_row(current_row, "Subtotal Pesanan", "", subtotal_pesanan, indent=1)
        current_row = write_row(current_row, "Harga Asli Produk", "", harga_asli, indent=2)
        current_row = write_row(current_row, "Total Diskon Produk", "", total_diskon, indent=2)
        current_row = write_row(current_row, "Jumlah Pengembalian Dana ke Pembeli", "", refund_pembeli, indent=2)
        current_row += 1  # Empty row
        
        # Voucher & Subsidi Shopee
        current_row = write_row(current_row, "Voucher & Subsidi Shopee", "", total_voucher_subsidi, bold=True, indent=1)
        current_row = write_row(current_row, "Diskon Produk dari Shopee", "", diskon_shopee, indent=2)
        current_row = write_row(current_row, "Voucher dari Penjual", "", voucher_penjual, indent=2)
        current_row = write_row(current_row, "Cashback Koin dari Penjual", "", cashback_penjual, indent=2)
        current_row += 1  # Empty row
        
        # 2. Total Pengeluaran
        current_row = write_row(current_row, "2. Total Pengeluaran", "", total_pengeluaran, bold=True)
        current_row = write_row(current_row, "Total Biaya Pengiriman", "", total_biaya_pengiriman, indent=1)
        current_row = write_row(current_row, "Ongkir Dibayar Pembeli", "", ongkir_pembeli, indent=2)
        current_row = write_row(current_row, "Diskon Ongkir Ditanggung Jasa Kirim", "", diskon_ongkir_3pl, indent=2)
        current_row = write_row(current_row, "Gratis Ongkir dari Shopee", "", gratis_ongkir_shopee, indent=2)
        current_row = write_row(current_row, "Ongkir yang Diteruskan oleh Shopee ke Jasa Kirim", "", ongkir_diteruskan, indent=2)
        current_row = write_row(current_row, "Ongkos Kirim Pengembalian Barang", "", ongkir_return, indent=2)
        current_row = write_row(current_row, "Pengembalian Biaya Kirim", "", pengembalian_biaya_kirim, indent=2)
        current_row = write_row(current_row, "Kembali ke Biaya Pengiriman Pengirim", "", kembali_biaya_pengirim, indent=2)
        current_row += 1  # Empty row
        
        # Biaya Admin & Layanan
        current_row = write_row(current_row, "Biaya Admin & Layanan", "", total_biaya_admin, bold=True, indent=1)
        current_row = write_row(current_row, "Biaya Komisi AMS", "", biaya_ams, indent=2)
        current_row = write_row(current_row, "Biaya Administrasi", "", biaya_admin, indent=2)
        current_row = write_row(current_row, "Biaya Layanan", "", biaya_layanan, indent=2)
        current_row = write_row(current_row, "Biaya Proses Pesanan", "", biaya_proses, indent=2)
        current_row = write_row(current_row, "Premi", "", premi, indent=2)
        current_row = write_row(current_row, "Biaya Program Hemat Biaya Kirim", "", biaya_hemat_kirim, indent=2)
        current_row = write_row(current_row, "Biaya Transaksi", "", biaya_transaksi, indent=2)
        current_row = write_row(current_row, "Biaya Kampanye", "", biaya_kampanye, indent=2)
        current_row += 2  # Empty rows
        
        # 3. Total yang Dilepas
        current_row = write_row(current_row, "3. Total yang Dilepas", "", total_dilepas, bold=True)
        current_row += 2  # Empty rows
        
        # Nilai Lainnya
        current_row = write_row(current_row, "Nilai Lainnya", "", "", bold=True)
        current_row = write_row(current_row, "Promo Gratis Ongkir dari Penjual", "", promo_gratis_ongkir)
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 50
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 20
        
        # Format numbers in column C as currency
        for row in ws_summary.iter_rows(min_row=1, max_row=current_row, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    cell.number_format = '#,##0'
        
        # --- SHEET 2: Income Detail (40+ Kolom) ---
        df_inc.to_excel(writer, sheet_name='Income', index=False)
        ws_income = writer.sheets['Income']
        
        # Auto-adjust column widths
        for column in ws_income.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_income.column_dimensions[column_letter].width = adjusted_width
        
        # --- SHEET 3: Service Fee Details ---
        df_srv.to_excel(writer, sheet_name='Service Fee Details', index=False)
        ws_srv = writer.sheets['Service Fee Details']
        
        for column in ws_srv.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_srv.column_dimensions[column_letter].width = adjusted_width
        
        # --- SHEET 4: Order Processing Fee ---
        df_prc.to_excel(writer, sheet_name='Order Processing Fee', index=False)
        ws_prc = writer.sheets['Order Processing Fee']
        
        for column in ws_prc.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_prc.column_dimensions[column_letter].width = adjusted_width
        
    return output.getvalue()

def to_wib(timestamp):
    """Konversi timestamp ke WIB (UTC+7)"""
    if not timestamp:
        return ""
    try:
        # Jika timestamp dalam detik
        if isinstance(timestamp, (int, float)):
            dt_utc = datetime.fromtimestamp(timestamp, pytz.UTC)  # ✅ Pastikan UTC
            dt_wib = dt_utc.astimezone(WIB)  # ✅ Gunakan WIB variable
            return dt_wib.strftime('%Y-%m-%d %H:%M:%S')
        return str(timestamp)
    except:
        return ""

def to_wib_date(timestamp):
    """Konversi timestamp ke format tanggal WIB saja"""
    if not timestamp:
        return ""
    try:
        if isinstance(timestamp, (int, float)):
            dt_utc = datetime.fromtimestamp(timestamp, pytz.UTC)  # ✅ Pastikan UTC
            dt_wib = dt_utc.astimezone(WIB)  # ✅ Gunakan WIB variable
            return dt_wib.strftime('%Y-%m-%d')
        return str(timestamp)
    except:
        return ""
        
# ===============================
# UI
# ===============================
st.title("🤖 Humanoid - Shopee API Integration")
st.caption(f"🕐 Server UTC: {datetime.now(pytz.UTC)} | WIB: {datetime.now(WIB)}")

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "1️⃣ Authorisasi",
    "2️⃣ Tukar Code → Token",
    "3️⃣ Order-all & Detail",
    "4️⃣ Income (Dana Dilepas)",
    "5️⃣ Data Iklan Keseluruhan",
    "6️⃣ Seller Conversion",
    "🕒 Performa Iklan Per Jam"
])

# ===============================
# TAB 1 — AUTHORISASI
# ===============================
with tab1:
    st.header("Generate Authorization URL")

    if st.button("🔐 Generate Authorization URL"):
        path = "/api/v2/shop/auth_partner"
        timestamp = int(time.time())
        sign = generate_sign_basic(path, timestamp)

        params = {
            "partner_id": PARTNER_ID,
            "timestamp": timestamp,
            "sign": sign,
            "redirect": REDIRECT_URL
        }

        auth_url = BASE_URL + path + "?" + urllib.parse.urlencode(params)

        st.success("Buka URL ini untuk authorize toko:")
        st.code(auth_url)

        st.info("""
1. Klik URL di atas
2. Login Seller Shopee
3. Approve App
4. Redirect dengan ?code=XXXX&shop_id=YYYY
""")

# ===============================
# TAB 2 — TUKAR CODE → TOKEN
# ===============================
with tab2:
    st.header("Tukar Code ke Access Token")

    #code = st.text_input("Masukkan code dari redirect")
    #shop_id_input = st.text_input("Masukkan shop_id dari redirect")
    code = st.text_input(
        "Masukkan code dari redirect",
        value=oauth_code if oauth_code else ""
    )

    shop_id_input = st.text_input(
        "Masukkan shop_id dari redirect",
        value=oauth_shop_id if oauth_shop_id else ""
    )
    shop_name_input = st.text_input("Beri Nama Toko (misal: Human)", "Human")

    if st.button("🔄 Tukar ke Access Token"):
        path = "/api/v2/auth/token/get"
        timestamp = int(time.time())
        sign = generate_sign_basic(path, timestamp)

        url = BASE_URL + path
        
        # 1. Parameter untuk URL (Common Parameters)
        params = {
            "partner_id": int(PARTNER_ID),
            "timestamp": timestamp,
            "sign": sign
        }
        
        # 2. Parameter untuk Body (JSON Payload)
        payload = {
            "code": code,
            "shop_id": int(shop_id_input),
            "partner_id": int(PARTNER_ID)
        }

        # Kirim dengan memisahkan params (URL) dan json (Body)
        data = requests.post(url, params=params, json=payload).json()

        st.subheader("Response Token")
        st.json(data)

        if "access_token" in data:
            save_token_to_db(
                shop_name_input,
                shop_id_input,
                data["access_token"],
                data["refresh_token"]
            )
            st.success(f"Token toko '{shop_name_input}' berhasil disimpan!")
            # Tambahkan st.rerun() agar Tab 3 langsung update
            st.rerun()
        else:
            st.error("Gagal mendapatkan access_token. Cek kembali respons di atas.")

# ===============================
# TAB 3 — ORDER-ALL & DETAIL
# ===============================
with tab3:
    st.header("📦 Tarik Order-all & Order Detail")

    shop_names = get_all_shops()
    if not shop_names:
        st.warning("Belum ada toko di database. Authorize dulu.")
    else:
        selected_shop = st.selectbox("Pilih Toko", shop_names, key="shop_order_all")
        
        # 🔴 PERBAIKAN: Pilihan range tanggal fleksibel
        st.subheader("📅 Pilih Periode")
        col_preset, col_custom = st.columns([1, 2])
        
        with col_preset:
            preset_option = st.selectbox(
                "Preset Range",
                ["Custom", "Hari Ini", "Kemarin", "7 Hari Terakhir", "30 Hari Terakhir", "Minggu Ini", "Minggu Lalu", "Bulan Ini", "Bulan Lalu"],
                key="preset_range"
            )
        
        # Set default dates berdasarkan preset
        # 🔴 PERBAIKAN: Gunakan dt.date.today() bukan datetime.date.today()
        today = datetime.now(WIB).date()
        
        if preset_option == "Hari Ini":
            default_start = today
            default_end = today
        elif preset_option == "Kemarin":
            default_start = today - dt.timedelta(days=1)
            default_end = today - dt.timedelta(days=1)
        elif preset_option == "7 Hari Terakhir":
            default_start = today - dt.timedelta(days=7)
            default_end = today
        elif preset_option == "30 Hari Terakhir":
            default_start = today - dt.timedelta(days=30)
            default_end = today
        elif preset_option == "Minggu Ini":
            # Senin minggu ini
            default_start = today - dt.timedelta(days=today.weekday())
            default_end = today
        elif preset_option == "Minggu Lalu":
            # Senin sampai Minggu minggu lalu
            default_start = today - dt.timedelta(days=today.weekday() + 7)
            default_end = today - dt.timedelta(days=today.weekday() + 1)
        elif preset_option == "Bulan Ini":
            default_start = today.replace(day=1)
            default_end = today
        elif preset_option == "Bulan Lalu":
            last_month_end = today.replace(day=1) - dt.timedelta(days=1)
            default_start = last_month_end.replace(day=1)
            default_end = last_month_end
        else:  # Custom
            default_start = today - dt.timedelta(days=7)
            default_end = today
        
        with col_custom:
            col_a, col_b = st.columns(2)
            with col_a:
                start_date = st.date_input("Tanggal Mulai", default_start, key="order_start_date")
            with col_b:
                end_date = st.date_input("Tanggal Akhir", default_end, key="order_end_date")

        # 🔴 PERBAIKAN: Konversi ke timestamp dengan timezone WIB
        # Buat datetime dengan timezone Asia/Jakarta
        start_dt_wib = WIB.localize(datetime.combine(start_date, dt_time.min))
        end_dt_wib = WIB.localize(datetime.combine(end_date, dt_time.max))
        
        # ✅ PERBAIKAN: Konversi ke UTC untuk API Shopee (API menggunakan UTC)
        time_from = int(start_dt_wib.astimezone(UTC).timestamp())
        time_to = int(end_dt_wib.astimezone(UTC).timestamp())

        
        # Debug untuk verifikasi
        st.caption(f"🕐 WIB: {start_dt_wib.strftime('%Y-%m-%d %H:%M')} s/d {end_dt_wib.strftime('%Y-%m-%d %H:%M')} | UTC Timestamp: {time_from} - {time_to}")

        if st.button("📥 Tarik Order-all", type="primary"):
            token_row = get_shop_token(selected_shop)
            ACTIVE_SHOP_ID = token_row["shop_id"]
            ACTIVE_ACCESS_TOKEN = token_row["access_token"]

            all_order_sns = []
            cursor = ""
            status_info = st.empty()
            
            # 1. LOOP LIST ORDER
            with st.spinner("Mengambil daftar pesanan..."):
                while True:
                    path = "/api/v2/order/get_order_list"
                    ts = int(time.time())
                    sign = generate_sign_full(path, ts, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)
                    params = {
                        "partner_id": PARTNER_ID, "timestamp": ts, "access_token": ACTIVE_ACCESS_TOKEN,
                        "shop_id": int(ACTIVE_SHOP_ID), "sign": sign,
                        "time_range_field": "create_time", "time_from": time_from, "time_to": time_to,
                        "page_size": 50, "cursor": cursor
                    }
                    res = requests.get(BASE_URL + path, params=params).json()
                    resp_data = res.get("response", {})
                    orders = resp_data.get("order_list", [])
                    for o in orders:
                        all_order_sns.append(o["order_sn"])
                    
                    if not resp_data.get("has_next_page"): 
                        break
                    cursor = resp_data.get("next_cursor")
                    status_info.info(f"📄 Mengambil daftar pesanan... ({len(all_order_sns)} order ditemukan)")

            if not all_order_sns:
                st.warning("⚠️ Tidak ada pesanan dalam periode yang dipilih.")
                st.stop()

            status_map_indo = {
                "UNPAID": "Belum Bayar",
                "READY_TO_SHIP": "Perlu Dikirim",
                "PROCESSED": "Sedang Diproses",
                "SHIPPED": "Dikirim",
                "TO_CONFIRM_RECEIVE": "Sedang Dikirim",
                "COMPLETED": "Selesai",
                "IN_CANCEL": "Pengajuan Batal",
                "CANCELLED": "Batal",
                "TO_RETURN": "Pengajuan Pengembalian"
            }

            cancel_reason_map = {
                "OUT_OF_STOCK": "Kehabisan Stok",
                "CUSTOMER_REQUEST": "Permintaan Pembeli",
                "UNDELIVERABLE_AREA": "Area Tidak Terjangkau",
                "COD_NOT_SUPPORTED": "COD Tidak Didukung",
                "SELLER_REQUEST": "Permintaan Penjual",
                "SYSTEM_CANCEL": "Dibatalkan Sistem"
            }

            # 2. DETAIL & FINANCE (ESCROW)
            rows = []
            progress_bar = st.progress(0)
            total_orders = len(all_order_sns)
            
            for i in range(0, total_orders, 50):
                batch_sns = all_order_sns[i:i+50]
                path2 = "/api/v2/order/get_order_detail"
                ts2 = int(time.time())
                sign2 = generate_sign_full(path2, ts2, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)
                
                # Semua field yang diperlukan
                opt_fields = ("buyer_username,recipient_address,estimated_shipping_fee,actual_shipping_fee,"
                             "pay_time,ship_by_date,order_status,cancel_reason,item_list,payment_method,"
                             "shipping_carrier,note,message,create_time,finish_time,tracking_number,"
                             "total_amount,pickup_done_time,return_status,arrange_shipment_date,cod")

                p2 = {
                    "partner_id": PARTNER_ID, "timestamp": ts2, "access_token": ACTIVE_ACCESS_TOKEN,
                    "shop_id": int(ACTIVE_SHOP_ID), "sign": sign2, "order_sn_list": ",".join(batch_sns),
                    "response_optional_fields": opt_fields
                }
                
                detail_res = requests.get(BASE_URL + path2, params=p2).json()
                orders_detail = detail_res.get("response", {}).get("order_list", [])

                for o in orders_detail:
                    order_sn = o.get("order_sn")
                    status_info.info(f"🔍 Memproses: {order_sn} ({len(rows)} baris)")
                    
                    time.sleep(0.2)  # Rate limiting
                    
                    # Ambil escrow detail untuk data finance
                    esc = get_escrow_detail(order_sn, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)
                    
                    # 🔴 PERBAIKAN: Handle struktur response escrow
                    if "response" in esc:
                        income_info = esc.get("response", {}).get("order_income", {})
                    else:
                        income_info = esc.get("order_income", {})
                    
                    addr = o.get("recipient_address", {}) or {}

                    # Terjemahkan Status
                    raw_status = o.get("order_status", "")
                    indo_status = status_map_indo.get(raw_status, raw_status)
                    
                    # Terjemahkan Alasan Batal
                    raw_reason = o.get("cancel_reason", "")
                    indo_reason = cancel_reason_map.get(raw_reason, raw_reason) if raw_reason else ""
                    
                    # 🔴 PERBAIKAN: Format tanggal ke WIB
                    ship_by_date = to_wib(o.get("ship_by_date"))
                    arrange_shipment_date = to_wib(o.get("arrange_shipment_date"))
                    create_time = to_wib(o.get("create_time"))
                    pay_time = to_wib(o.get("pay_time"))
                    finish_time = to_wib(o.get("finish_time"))
                    
                    # Determine pickup/counter
                    pickup_done_time = o.get("pickup_done_time")
                    antar_counter = "Pick-up" if pickup_done_time else "Counter"
                    
                    # Items loop
                    items = o.get("item_list", [])
                    for item in items:
                        qty = item.get("model_quantity_purchased", 0) or item.get("quantity_purchased", 0) or 0
                        price = item.get("model_discounted_price", 0) or item.get("discounted_price", 0) or 0
                        original_price = item.get("model_original_price", 0) or item.get("original_price", 0) or 0
                        weight = item.get("weight", 0) or 0
                        
                        # 🔴 PERBAIKAN: Mapping lengkap 42 kolom
                        row = {
                            "No. Pesanan": order_sn,
                            "Status Pesanan": indo_status,
                            "Alasan Pembatalan": indo_reason,
                            "Status Pembatalan/ Pengembalian": o.get("return_status", ""),
                            "No. Resi": o.get("tracking_number", ""),
                            "Opsi Pengiriman": o.get("shipping_carrier", ""),
                            "Antar ke counter/ pick-up": antar_counter,
                            "Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)": ship_by_date,
                            "Waktu Pengiriman Diatur": arrange_shipment_date,
                            "Waktu Pesanan Dibuat": create_time,
                            "Waktu Pembayaran Dilakukan": pay_time,
                            "Metode Pembayaran": o.get("payment_method", ""),
                            "SKU Induk": item.get("item_sku", ""),
                            "Nama Produk": item.get("item_name", ""),
                            "Nomor Referensi SKU": item.get("model_sku", ""),
                            "Nama Variasi": item.get("model_name", ""),
                            "Harga Awal": original_price,
                            "Harga Setelah Diskon": price,
                            "Jumlah": qty,
                            "Returned quantity": item.get("is_return_item", False),
                            "Total Harga Produk": price * qty,
                            
                            # Finance dari Escrow
                            "Total Diskon": (income_info.get("seller_discount", 0) or 0) + (income_info.get("shopee_discount", 0) or 0),
                            "Diskon Dari Penjual": income_info.get("seller_discount", 0) or 0,
                            "Diskon Dari Shopee": income_info.get("shopee_discount", 0) or 0,
                            "Berat Produk": weight,
                            "Jumlah Produk di Pesan": qty,
                            "Total Berat": weight * qty,
                            "Voucher Ditanggung Penjual": income_info.get("voucher_from_seller", 0) or 0,
                            "Cashback Koin": income_info.get("coins", 0) or income_info.get("seller_coin_cash_back", 0) or 0,
                            "Voucher Ditanggung Shopee": income_info.get("voucher_from_shopee", 0) or 0,
                            "Paket Diskon": (income_info.get("bundle_discount_from_seller", 0) or 0) + (income_info.get("bundle_discount_from_shopee", 0) or 0),
                            "Paket Diskon (Diskon dari Shopee)": income_info.get("bundle_discount_from_shopee", 0) or 0,
                            "Paket Diskon (Diskon dari Penjual)": income_info.get("bundle_discount_from_seller", 0) or 0,
                            "Potongan Koin Shopee": income_info.get("coins", 0) or income_info.get("seller_coin_cash_back", 0) or 0,
                            "Diskon Kartu Kredit": income_info.get("credit_card_promotion", 0) or 0,
                            "Ongkos Kirim Dibayar oleh Pembeli": o.get("actual_shipping_fee", 0) or income_info.get("buyer_paid_shipping_fee", 0) or 0,
                            "Estimasi Potongan Biaya Pengiriman": income_info.get("shopee_shipping_rebate", 0) or 0,
                            "Ongkos Kirim Pengembalian Barang": income_info.get("reverse_shipping_fee", 0) or 0,
                            "Total Pembayaran": o.get("total_amount", 0),
                            "Perkiraan Ongkos Kirim": o.get("estimated_shipping_fee", 0),
                            
                            "Catatan dari Pembeli": o.get("message", ""),
                            "Catatan": o.get("note", ""),
                            "Username (Pembeli)": o.get("buyer_username", ""),
                            "Nama Penerima": addr.get("name", ""),
                            "No. Telepon": addr.get("phone", ""),
                            "Alamat Pengiriman": addr.get("full_address", ""),
                            "Kota/Kabupaten": addr.get("city", ""),
                            "Provinsi": addr.get("state", ""),
                            "Waktu Pesanan Selesai": finish_time
                        }
                        
                        rows.append(row)

                progress_pct = min((i + 50) / total_orders, 1.0)
                progress_bar.progress(progress_pct)

            if not rows:
                st.error("❌ Tidak ada data yang berhasil diproses.")
                st.stop()

            df = pd.DataFrame(rows)
            
            # 🔴 PERBAIKAN: Urutkan kolom sesuai permintaan
            desired_columns = [
                "No. Pesanan", "Status Pesanan", "Alasan Pembatalan", "Status Pembatalan/ Pengembalian",
                "No. Resi", "Opsi Pengiriman", "Antar ke counter/ pick-up",
                "Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)", "Waktu Pengiriman Diatur",
                "Waktu Pesanan Dibuat", "Waktu Pembayaran Dilakukan", "Metode Pembayaran",
                "SKU Induk", "Nama Produk", "Nomor Referensi SKU", "Nama Variasi",
                "Harga Awal", "Harga Setelah Diskon", "Jumlah", "Returned quantity", "Total Harga Produk",
                "Total Diskon", "Diskon Dari Penjual", "Diskon Dari Shopee", "Berat Produk",
                "Jumlah Produk di Pesan", "Total Berat", "Voucher Ditanggung Penjual", "Cashback Koin",
                "Voucher Ditanggung Shopee", "Paket Diskon", "Paket Diskon (Diskon dari Shopee)",
                "Paket Diskon (Diskon dari Penjual)", "Potongan Koin Shopee", "Diskon Kartu Kredit",
                "Ongkos Kirim Dibayar oleh Pembeli", "Estimasi Potongan Biaya Pengiriman",
                "Ongkos Kirim Pengembalian Barang", "Total Pembayaran", "Perkiraan Ongkos Kirim",
                "Catatan dari Pembeli", "Catatan", "Username (Pembeli)", "Nama Penerima",
                "No. Telepon", "Alamat Pengiriman", "Kota/Kabupaten", "Provinsi", "Waktu Pesanan Selesai"
            ]
            
            # Filter hanya kolom yang ada
            available_columns = [col for col in desired_columns if col in df.columns]
            df = df[available_columns]

            # Generate Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                df.to_excel(writer, index=False, sheet_name="Order-All")
                # Auto-adjust columns
                worksheet = writer.sheets["Order-All"]
                for i, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.set_column(i, i, min(max_len, 50))
            
            excel_bytes = output.getvalue()
            
            range_str = f"{start_date} s/d {end_date}"
            save_report_to_db(selected_shop, range_str, excel_bytes)
            
            st.success(f"✅ Berhasil! {len(df)} baris produk dari {len(all_order_sns)} pesanan.")
            
            # Preview
            st.subheader("📋 Preview Data")
            st.dataframe(df.head(20), use_container_width=True)

            st.download_button(
                "⬇️ Download Order-all Excel",
                excel_bytes,
                f"Order_{selected_shop}_{start_date}_{end_date}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.divider()
        st.subheader("📜 Riwayat Laporan (Database)")
        history = get_report_history(selected_shop)
        if not history:
            st.write("Belum ada riwayat laporan untuk toko ini.")
        else:
            for item in history:
                col1, col2, col3 = st.columns([3, 3, 2])
                col1.write(f"📅 {item['date_range']}")
                col2.write(f"⏰ {item['created_at'][:19]}")
                report_bytes = base64.b64decode(item['csv_content'])
                col3.download_button(
                    label="💾 Download",
                    data=report_bytes,
                    file_name=f"Order_{selected_shop}_{item['created_at'][:10]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=item['id']
                )

# ==========================================
# TAB 4: INCOME (DANA DILEPASKAN)
# ==========================================
with tab4:
    st.header("💰 Laporan Income (Dana Dilepaskan)")
    st.info("Data ditarik berdasarkan tanggal dana masuk ke saldo (Released) dalam timezone WIB.")

    if not shop_names:
        st.warning("Belum ada toko.")
    else:
        selected_shop_inc = st.selectbox("Pilih Toko untuk Income", shop_names, key="shop_income_ui")
        
        # 🔴 PERBAIKAN: Pilihan range tanggal fleksibel
        st.subheader("📅 Pilih Periode")
        col_preset_inc, col_custom_inc = st.columns([1, 2])
        
        with col_preset_inc:
            preset_inc = st.selectbox(
                "Preset Range",
                ["Custom", "Hari Ini", "Kemarin", "7 Hari Terakhir", "30 Hari Terakhir", 
                 "Minggu Ini", "Minggu Lalu", "Bulan Ini", "Bulan Lalu"],
                key="preset_income"
            )
        
        # 🔴 PERBAIKAN: Gunakan dt.date.today()
        today = datetime.now(WIB).date()
        
        if preset_inc == "Hari Ini":
            default_start_inc = today
            default_end_inc = today
        elif preset_inc == "Kemarin":
            default_start_inc = today - dt.timedelta(days=1)
            default_end_inc = today - dt.timedelta(days=1)
        elif preset_inc == "7 Hari Terakhir":
            default_start_inc = today - dt.timedelta(days=7)
            default_end_inc = today
        elif preset_inc == "30 Hari Terakhir":
            default_start_inc = today - dt.timedelta(days=30)
            default_end_inc = today
        elif preset_inc == "Minggu Ini":
            default_start_inc = today - dt.timedelta(days=today.weekday())
            default_end_inc = today
        elif preset_inc == "Minggu Lalu":
            default_start_inc = today - dt.timedelta(days=today.weekday() + 7)
            default_end_inc = today - dt.timedelta(days=today.weekday() + 1)
        elif preset_inc == "Bulan Ini":
            default_start_inc = today.replace(day=1)
            default_end_inc = today
        elif preset_inc == "Bulan Lalu":
            last_month_end = today.replace(day=1) - dt.timedelta(days=1)
            default_start_inc = last_month_end.replace(day=1)
            default_end_inc = last_month_end
        else:
            default_start_inc = today - dt.timedelta(days=7)
            default_end_inc = today
        
        with col_custom_inc:
            col_i1, col_i2 = st.columns(2)
            with col_i1:
                start_inc = st.date_input("Dari Tanggal", default_start_inc, key="dt_start_inc")
            with col_i2:
                end_inc = st.date_input("Sampai Tanggal", default_end_inc, key="dt_end_inc")

        if st.button("📊 Generate Laporan Income", type="primary"):
            token_row = get_shop_token(selected_shop_inc)
            ACTIVE_SHOP_ID = token_row["shop_id"]
            ACTIVE_ACCESS_TOKEN = token_row["access_token"]
            
            prog_bar = st.progress(0)
            status_text = st.empty()
            
            all_sn_list = []
            income_rows = []
            service_rows = []
            
            # 🔴 PERBAIKAN: Konversi waktu dengan timezone WIB
            start_dt_wib = WIB.localize(datetime.combine(start_inc, dt_time.min))
            end_dt_wib = WIB.localize(datetime.combine(end_inc, dt_time.max))
            
            # ✅ PERBAIKAN: Konversi ke UTC timestamp untuk API Shopee
            time_from = int(start_dt_wib.astimezone(UTC).timestamp())
            time_to = int(end_dt_wib.astimezone(UTC).timestamp())

            
            status_text.info(f"🔍 Mencari dana dilepaskan: {start_inc} s/d {end_inc} (WIB)")

            # 1. AMBIL DAFTAR PESANAN CAIR
            path_esc = "/api/v2/payment/get_escrow_list"
            page_no = 1
            
            with st.spinner("Mengambil daftar escrow..."):
                while True:
                    ts = int(time.time())
                    sign = generate_sign_full(path_esc, ts, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)
                    
                    params = {
                        "partner_id": PARTNER_ID, 
                        "timestamp": ts, 
                        "access_token": ACTIVE_ACCESS_TOKEN,
                        "shop_id": int(ACTIVE_SHOP_ID), 
                        "sign": sign,
                        "release_time_from": time_from, 
                        "release_time_to": time_to,
                        "page_size": 50,
                        "page_no": page_no
                    }
                    
                    try:
                        response = requests.get(BASE_URL + path_esc, params=params, timeout=30)
                        res = response.json()
                        
                        if res.get("error"):
                            st.error(f"❌ API Error: {res.get('message')}")
                            break
                        
                        escrow_data = res.get("response", {}).get("escrow_list", [])
                        
                        if escrow_data:
                            for item in escrow_data:
                                all_sn_list.append(item.get("order_sn"))
                        
                        has_more = res.get("response", {}).get("more", False)
                        status_text.info(f"📄 Halaman {page_no}: {len(all_sn_list)} order ditemukan")
                        
                        if not has_more or not escrow_data:
                            break
                        page_no += 1
                        
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        break

            # 2. PROSES DETAIL
            if not all_sn_list:
                st.error(f"❌ Tidak ada dana cair pada periode {start_inc} s/d {end_inc}.")
            else:
                st.success(f"✅ Ditemukan {len(all_sn_list)} order dengan dana dilepaskan")
                
                for idx, sn in enumerate(all_sn_list):
                    progress_pct = (idx + 1) / len(all_sn_list)
                    prog_bar.progress(progress_pct)
                    status_text.text(f"[{idx+1}/{len(all_sn_list)}] Memproses: {sn}")
                    
                    try:
                        # Ambil Escrow Detail
                        esc_res = get_escrow_detail(sn, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)
                        
                        # Handle struktur response
                        if "response" in esc_res:
                            oi = esc_res.get("response", {}).get("order_income", {})
                            buyer_username_esc = esc_res.get("response", {}).get("buyer_user_name", "")
                        else:
                            oi = esc_res.get("order_income", {})
                            buyer_username_esc = esc_res.get("buyer_user_name", "")
                        
                        if not oi:
                            st.warning(f"⚠️ {sn}: Tidak ada data income")
                            continue
                        
                        # Ambil Order Detail
                        path_dtl = "/api/v2/order/get_order_detail"
                        ts_dtl = int(time.time())
                        sign_dtl = generate_sign_full(path_dtl, ts_dtl, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)
                        p_dtl = {
                            "partner_id": PARTNER_ID, 
                            "timestamp": ts_dtl, 
                            "access_token": ACTIVE_ACCESS_TOKEN,
                            "shop_id": int(ACTIVE_SHOP_ID), 
                            "sign": sign_dtl, 
                            "order_sn_list": sn,
                            "response_optional_fields": "item_list,buyer_username,create_time,shipping_carrier,payment_method"
                        }
                        
                        ord_res = requests.get(BASE_URL + path_dtl, params=p_dtl, timeout=30).json()
                        
                        if "response" in ord_res:
                            ord_list = ord_res.get("response", {}).get("order_list", [])
                        else:
                            ord_list = ord_res.get("order_list", [])
                        
                        ord_dtl = ord_list[0] if ord_list else {}
                        
                        # 🔴 PERBAIKAN: Format tanggal ke WIB
                        release_ts = oi.get("escrow_release_time")
                        release_dt = to_wib_date(release_ts)
                        
                        create_ts = ord_dtl.get("create_time") if ord_dtl else None
                        create_dt = to_wib(create_ts)
                        
                        # Hitung diskon
                        seller_disc = oi.get("seller_discount", 0) or 0
                        shopee_disc = oi.get("shopee_discount", 0) or 0
                        total_disc = (seller_disc + shopee_disc) * -1
                        
                        final_buyer = ord_dtl.get("buyer_username") if ord_dtl else buyer_username_esc
                        
                        # --- MAPPING 41 KOLOM ---
                        row_data = {
                            "No.": idx + 1,
                            "No. Pesanan": sn,
                            "No. Pengajuan": "",
                            "Username (Pembeli)": final_buyer,
                            "Waktu Pesanan Dibuat": create_dt,
                            "Metode pembayaran pembeli": ord_dtl.get("payment_method", "") if ord_dtl else oi.get("buyer_payment_method", ""),
                            "Tanggal Dana Dilepaskan": release_dt,
                            "Harga Asli Produk": oi.get("original_cost_of_goods_sold", 0) or oi.get("original_price", 0) or 0,
                            "Total Diskon Produk": total_disc,
                            "Jumlah Pengembalian Dana ke Pembeli": oi.get("seller_return_refund", 0) or 0,
                            "Diskon Produk dari Shopee": shopee_disc,
                            "Voucher dari Penjual": oi.get("voucher_from_seller", 0) or 0,
                            "Cashback Koin dari Penjual": oi.get("seller_coin_cash_back", 0) or 0,
                            "Ongkir Dibayar Pembeli": oi.get("buyer_paid_shipping_fee", 0) or 0,
                            "Diskon Ongkir Ditanggung Jasa Kirim": oi.get("shipping_fee_discount_from_3pl", 0) or 0,
                            "Gratis Ongkir dari Shopee": oi.get("shopee_shipping_rebate", 0) or 0,
                            "Ongkir yang Diteruskan oleh Shopee ke Jasa Kirim": oi.get("actual_shipping_fee", 0) or 0,
                            "Ongkos Kirim Pengembalian Barang": oi.get("reverse_shipping_fee", 0) or 0,
                            "Kembali ke Biaya Pengiriman Pengirim": 0,
                            "Pengembalian Biaya Kirim": 0,
                            "Biaya Komisi AMS": oi.get("order_ams_commission_fee", 0) or 0,
                            "Biaya Administrasi": oi.get("commission_fee", 0) or 0,
                            "Biaya Layanan": oi.get("service_fee", 0) or 0,
                            "Biaya Proses Pesanan": oi.get("seller_transaction_fee", 0) or oi.get("seller_order_processing_fee", 0) or 0,
                            "Premi": oi.get("delivery_seller_protection_fee_premium_amount", 0) or 0,
                            "Biaya Program Hemat Biaya Kirim": 0,
                            "Biaya Transaksi": oi.get("seller_transaction_fee", 0) or 0,
                            "Biaya Kampanye": oi.get("campaign_fee", 0) or 0,
                            "Bea Masuk, PPN & PPh": (oi.get("escrow_tax", 0) or 0) + (oi.get("withholding_tax", 0) or 0),
                            "Total Penghasilan": oi.get("escrow_amount", 0) or 0,
                            "Kode Voucher": ",".join(oi.get("seller_voucher_code", [])) if oi.get("seller_voucher_code") else "",
                            "Kompensasi": oi.get("seller_lost_compensation", 0) or 0,
                            "Promo Gratis Ongkir dari Penjual": oi.get("seller_shipping_discount", 0) or 0,
                            "Jasa Kirim": ord_dtl.get("shipping_carrier", "") if ord_dtl else "",
                            "Nama Kurir": ord_dtl.get("shipping_carrier", "") if ord_dtl else "",
                            "Pengembalian Dana ke Pembeli": oi.get("seller_return_refund", 0) or 0,
                            "Pro-rata Koin yang Ditukarkan untuk Pengembalian Barang": oi.get("prorated_coins_value_offset_return_items", 0) or 0,
                            "Pro-rata Voucher Shopee untuk Pengembalian Barang": oi.get("prorated_shopee_voucher_offset_return_items", 0) or 0,
                            "Pro-rated Bank Payment Channel Promotion for return refund Items": oi.get("prorated_payment_channel_promo_bank_offset_return_items", 0) or 0,
                            "Pro-rated Shopee Payment Channel Promotion for return refund Items": oi.get("prorated_payment_channel_promo_shopee_offset_return_items", 0) or 0
                        }
                        
                        income_rows.append(row_data)
                        
                        service_rows.append({
                            "No.": idx + 1, 
                            "No. Pesanan": sn, 
                            "Biaya Layanan Gratis Ongkir XTRA": oi.get("service_fee", 0) or 0
                        })
                        
                    except Exception as e:
                        st.error(f"❌ Error {sn}: {str(e)}")
                        continue

                status_text.empty()
                
                if income_rows:
                    df_inc = pd.DataFrame(income_rows)
                    df_srv = pd.DataFrame(service_rows)
                    
                    st.subheader("📋 Preview Data Income")
                    st.dataframe(df_inc.head(10), use_container_width=True)
                    
                    # Generate Excel
                    df_prc = df_inc[["No.", "No. Pesanan", "Biaya Proses Pesanan"]].copy()
                    df_prc["View By"] = "Order"
                    
                    excel_file = create_income_excel(df_inc, df_srv, df_prc, selected_shop_inc, str(start_inc), str(end_inc))
                    
                    total_income = df_inc["Total Penghasilan"].sum()
                    range_inc_str = f"{start_inc} s/d {end_inc}"

                    save_report_to_db(selected_shop_inc, f"INCOME {range_inc_str}", excel_file)
                    
                    st.success(f"✅ Berhasil! {len(df_inc)} data, Total: Rp {total_income:,.0f}")
                    
                    col_dl, col_stat = st.columns([1, 2])
                    with col_dl:
                        st.download_button(
                            label="📥 Download Excel",
                            data=excel_file,
                            file_name=f"Income_{selected_shop_inc}_{start_inc}_{end_inc}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    with col_stat:
                        st.info(f"Periode: {range_inc_str} (WIB) | Total Order: {len(df_inc)}")
                else:
                    st.error("❌ Tidak ada data yang berhasil diproses.")

        st.divider()
        st.subheader("📜 Riwayat Laporan Income")
        
        history_inc = get_report_history(selected_shop_inc)
        
        if not history_inc:
            st.write("Belum ada riwayat.")
        else:
            for item in history_inc:
                if not item["date_range"].startswith("INCOME"):
                    continue
            
                col1, col2, col3 = st.columns([3, 3, 2])
                col1.write(f"📅 {item['date_range'].replace('INCOME ', '')}")
                col2.write(f"⏰ {item['created_at'][:19]}")
                col3.download_button(
                    label="💾 Download",
                    data=item["csv_content"],
                    file_name=f"Income_{selected_shop_inc}_{item['created_at'][:10]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"inc_{item['id']}"
                )

        
with tab5:
    st.header("📢 Data Iklan Keseluruhan (Shopee Ads)")
    st.info("Mengambil data iklan menggunakan API Product Level dengan timezone WIB (UTC+7).")

    if not shop_names:
        st.warning("Belum ada toko.")
    else:
        selected_shop_ads = st.selectbox("Pilih Toko untuk Iklan", shop_names, key="shop_ads")

        # 🔴 PERBAIKAN: Pilihan range tanggal fleksibel
        st.subheader("📅 Pilih Periode")
        col_preset, col_custom = st.columns([1, 2])
        
        with col_preset:
            preset_ads = st.selectbox(
                "Preset Range",
                ["Custom", "Hari Ini", "Kemarin", "7 Hari Terakhir", "Minggu Ini", 
                 "Minggu Lalu", "Bulan Ini", "Bulan Lalu", "30 Hari Terakhir"],
                key="preset_ads"
            )
        
        # Set default dates berdasarkan preset
        today = datetime.now(WIB).date()
        
        if preset_ads == "Hari Ini":
            default_start = today
            default_end = today
        elif preset_ads == "Kemarin":
            default_start = today - dt.timedelta(days=1)
            default_end = today - dt.timedelta(days=1)
        elif preset_ads == "7 Hari Terakhir":
            default_start = today - dt.timedelta(days=7)
            default_end = today - dt.timedelta(days=1)  # Sampai kemarin
        elif preset_ads == "30 Hari Terakhir":
            default_start = today - dt.timedelta(days=30)
            default_end = today - dt.timedelta(days=1)
        elif preset_ads == "Minggu Ini":
            # Senin minggu ini sampai hari ini/kemarin
            default_start = today - dt.timedelta(days=today.weekday())
            default_end = today - dt.timedelta(days=1) if today.weekday() > 0 else today
        elif preset_ads == "Minggu Lalu":
            # Senin sampai Minggu minggu lalu
            default_start = today - dt.timedelta(days=today.weekday() + 7)
            default_end = today - dt.timedelta(days=today.weekday() + 1)
        elif preset_ads == "Bulan Ini":
            default_start = today.replace(day=1)
            default_end = today - dt.timedelta(days=1) if today.day > 1 else today
        elif preset_ads == "Bulan Lalu":
            last_month_end = today.replace(day=1) - dt.timedelta(days=1)
            default_start = last_month_end.replace(day=1)
            default_end = last_month_end
        else:  # Custom
            default_start = today - dt.timedelta(days=7)
            default_end = today - dt.timedelta(days=1)

        with col_custom:
            col_ad1, col_ad2 = st.columns(2)
            with col_ad1:
                start_ads = st.date_input("Dari Tanggal", default_start, key="s_ads")
            with col_ad2:
                end_ads = st.date_input("Sampai Tanggal", default_end, key="e_ads")

        # 🔴 PERBAIKAN: Validasi tanggal
        if start_ads > end_ads:
            st.error("❌ Tanggal mulai tidak boleh lebih besar dari tanggal akhir!")
            st.stop()

        # 🔴 PERBAIKAN: Info timezone
        st.caption(f"🕐 Periode: {start_ads} s/d {end_ads} (WIB - UTC+7)")
        
        # Konversi ke format string untuk API (DD-MM-YYYY)
        s_date_str = start_ads.strftime("%d-%m-%Y")
        e_date_str = end_ads.strftime("%d-%m-%Y")


        # Mapping untuk display
        status_map = {
            "ongoing": "Berjalan",
            "paused": "Dihentikan Sementara",
            "ended": "Selesai",
            "scheduled": "Dijadwalkan",
            "deleted": "Dihapus",
            "unknown": "Tidak Diketahui"
        }
        
        placement_map = {
            "all": "Semua Penempatan",
            "search": "Iklan Pencarian",
            "discovery": "Iklan Produk Serupa",
            "shop": "Iklan Toko",
            "other": "Lainnya"
        }
        
        bidding_map = {
            "auto": "GMV Max Auto",
            "manual": "Manual",
            "": "-",
            None: "-"
        }

        if st.button("📊 Tarik Data Iklan", type="primary"):
            token_row = get_shop_token(selected_shop_ads)
            ACTIVE_SHOP_ID = token_row["shop_id"]
            ACTIVE_ACCESS_TOKEN = token_row["access_token"]

            # Progress indicators
            prog_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.info("🔍 Mengambil daftar campaign...")

            # 1. AMBIL SEMUA CAMPAIGN ID
            path_list = "/api/v2/ads/get_product_level_campaign_id_list"
            ts = int(time.time())
            params_base = {
                "partner_id": int(PARTNER_ID),
                "timestamp": ts,
                "access_token": ACTIVE_ACCESS_TOKEN,
                "shop_id": int(ACTIVE_SHOP_ID),
                "sign": generate_sign_full(path_list, ts, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID),
                "ad_type": "all",
                "offset": 0,
                "limit": 5000
            }
            
            try:
                res_list = requests.get(BASE_URL + path_list, params=params_base, timeout=30).json()
                
                if res_list.get("error"):
                    st.error(f"❌ API Error (Campaign List): {res_list.get('message')}")
                    st.stop()
                
                campaign_entries = res_list.get("response", {}).get("campaign_list", [])
            except Exception as e:
                st.error(f"❌ Error fetching campaign list: {str(e)}")
                st.stop()
            
            if not campaign_entries:
                st.warning("⚠️ Tidak ada kampanye ditemukan untuk toko ini.")
                st.stop()
            
            st.success(f"✅ Ditemukan {len(campaign_entries)} campaign")
            
            # Prepare batches
            all_ids = [str(c["campaign_id"]) for c in campaign_entries]
            batch_size = 100
            id_batches = [all_ids[i:i + batch_size] for i in range(0, len(all_ids), batch_size)]
            
            final_results = []
            total_batches = len(id_batches)

            for idx, batch in enumerate(id_batches):
                batch_num = idx + 1
                status_text.info(f"📊 Memproses batch {batch_num}/{total_batches} ({len(batch)} campaign)...")
                prog_bar.progress(min(idx / total_batches, 0.95))
                
                ids_str = ",".join(batch)
                
                try:
                    # 2. AMBIL SETTING INFO
                    path_set = "/api/v2/ads/get_product_level_campaign_setting_info"
                    ts_set = int(time.time())
                    params_set = {
                        "partner_id": int(PARTNER_ID),
                        "timestamp": ts_set,
                        "access_token": ACTIVE_ACCESS_TOKEN,
                        "shop_id": int(ACTIVE_SHOP_ID),
                        "sign": generate_sign_full(path_set, ts_set, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID),
                        "info_type_list": "1,2,3,4",
                        "campaign_id_list": ids_str
                    }
                    
                    res_set = requests.get(BASE_URL + path_set, params=params_set, timeout=30).json()
                    
                    if res_set.get("error"):
                        st.warning(f"⚠️ Error settings batch {batch_num}: {res_set.get('message')}")
                        continue
                    
                    settings_list = res_set.get("response", {}).get("campaign_list", [])
                    settings_map = {str(s["campaign_id"]): s for s in settings_list}

                    # 3. AMBIL PERFORMANCE HARIAN
                    path_perf = "/api/v2/ads/get_product_campaign_daily_performance"
                    ts_perf = int(time.time())
                    params_perf = {
                        "partner_id": int(PARTNER_ID),
                        "timestamp": ts_perf,
                        "access_token": ACTIVE_ACCESS_TOKEN,
                        "shop_id": int(ACTIVE_SHOP_ID),
                        "sign": generate_sign_full(path_perf, ts_perf, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID),
                        "start_date": s_date_str,
                        "end_date": e_date_str,
                        "campaign_id_list": ids_str
                    }
                    
                    res_perf = requests.get(BASE_URL + path_perf, params=params_perf, timeout=30).json()
                    
                    if res_perf.get("error"):
                        st.warning(f"⚠️ Error performance batch {batch_num}: {res_perf.get('message')}")
                        continue
                    
                    perf_list = res_perf.get("response", {}).get("campaign_list", [])

                    # 4. MERGE & PROCESS DATA
                    for p_data in perf_list:
                        try:
                            cid = str(p_data.get("campaign_id", ""))
                            if not cid:
                                continue
                                
                            s_info = settings_map.get(cid, {})
                            common = s_info.get("common_info", {})
                            
                            # Skip jika tidak ada common_info
                            if not common:
                                continue
                            
                            # Status dengan fallback
                            raw_status = common.get("campaign_status", "unknown")
                            status_indo = status_map.get(raw_status, raw_status)

                            # if status_indo != "Berjalan":
                            #     continue
                            
                            # Mode Bidding
                            bidding_method = common.get("bidding_method", "")
                            mode_bidding = bidding_map.get(bidding_method, bidding_method if bidding_method else "-")
                            
                            # Penempatan
                            placement = common.get("campaign_placement", "all")
                            placement_indo = placement_map.get(placement, placement)
                            
                            # Produk Info
                            item_ids = common.get("item_id_list", [])
                            kode_produk = str(item_ids[0]) if item_ids else "-"
                            
                            # Tanggal dengan WIB
                            start_timestamp = common.get("start_time")
                            end_timestamp = common.get("end_time")
                            
                            if start_timestamp:
                                start_dt_wib = datetime.fromtimestamp(start_timestamp, pytz.UTC).astimezone(pytz.timezone('Asia/Jakarta'))
                                tanggal_mulai = start_dt_wib.strftime('%d/%m/%Y %H:%M:%S')
                            else:
                                tanggal_mulai = start_ads.strftime('%d/%m/%Y %H:%M:%S')
                            
                            if end_timestamp and end_timestamp != 0:
                                end_dt_wib = datetime.fromtimestamp(end_timestamp, pytz.UTC).astimezone(pytz.timezone('Asia/Jakarta'))
                                tanggal_selesai = end_dt_wib.strftime('%Y-%m-%d')
                            else:
                                tanggal_selesai = "Tidak Terbatas"
                            
                            # Agregasi Metrik
                            m_list = p_data.get("metrics_list", [])
                            
                            if not m_list:
                                # Jika tidak ada metrics, skip atau isi dengan 0
                                t_imp = t_cli = t_exp = t_gmv = t_ord = t_sold = d_gmv = d_ord = d_sold = 0
                            else:
                                t_imp = sum(m.get("impression", 0) or 0 for m in m_list)
                                t_cli = sum(m.get("clicks", 0) or 0 for m in m_list)
                                t_exp = sum(m.get("expense", 0) or 0 for m in m_list)
                                t_gmv = sum(m.get("broad_gmv", 0) or 0 for m in m_list)
                                t_ord = sum(m.get("broad_order", 0) or 0 for m in m_list)
                                t_sold = sum(m.get("broad_order_amount", 0) or 0 for m in m_list)
                                d_gmv = sum(m.get("direct_gmv", 0) or 0 for m in m_list)
                                d_ord = sum(m.get("direct_order", 0) or 0 for m in m_list)
                                d_sold = sum(m.get("direct_order_amount", 0) or 0 for m in m_list)
                            
                            # Kalkulasi Rasio (hindari division by zero)
                            ctr = (t_cli / t_imp * 100) if t_imp > 0 else 0
                            cvr = (t_ord / t_cli * 100) if t_cli > 0 else 0
                            cvr_direct = (d_ord / t_cli * 100) if t_cli > 0 else 0
                            acos = (t_exp / t_gmv * 100) if t_gmv > 0 else 0
                            acos_direct = (t_exp / d_gmv * 100) if d_gmv > 0 else 0
                            roas = (t_gmv / t_exp) if t_exp > 0 else 0
                            roas_direct = (d_gmv / t_exp) if t_exp > 0 else 0
                            cpc = (t_exp / t_cli) if t_cli > 0 else 0
                            
                            # Biaya per konversi
                            biaya_per_konversi = (t_exp / t_ord) if t_ord > 0 else 0
                            biaya_per_konversi_direct = (t_exp / d_ord) if d_ord > 0 else 0
                            
                            # 🔴 PERBAIKAN: Mapping lengkap 29 kolom
                            row_data = {
                                "Urutan": len(final_results) + 1,
                                "Nama Iklan": common.get("ad_name", p_data.get("ad_name", "-")),
                                "Status": status_indo,
                                "Jenis Iklan": "Iklan Produk",
                                "Kode Produk": kode_produk,
                                "Tampilan Iklan": "-" if t_imp == 0 else t_imp,
                                "Mode Bidding": mode_bidding,
                                "Penempatan Iklan": placement_indo,
                                "Tanggal Mulai": tanggal_mulai,
                                "Tanggal Selesai": tanggal_selesai,
                                "Dilihat": t_imp,
                                "Jumlah Klik": t_cli,
                                "Persentase Klik": f"{ctr:.2f}%",
                                "Konversi": t_ord,
                                "Konversi Langsung": d_ord,
                                "Tingkat konversi": f"{cvr:.2f}%",
                                "Tingkat Konversi Langsung": f"{cvr_direct:.2f}%",
                                "Biaya per Konversi": f"{biaya_per_konversi:.2f}",
                                "Biaya per Konversi Langsung": f"{biaya_per_konversi_direct:.2f}",
                                "Produk Terjual": t_sold,
                                "Terjual Langsung": d_sold,
                                "Omzet Penjualan": round(t_gmv, 0),
                                "Penjualan Langsung (GMV Langsung)": round(d_gmv, 0),
                                "Biaya": round(t_exp, 0),
                                "Efektifitas Iklan": f"{roas:.2f}",
                                "Efektivitas Langsung": f"{roas_direct:.2f}",
                                "Persentase Biaya Iklan terhadap Penjualan dari Iklan (ACOS)": f"{acos:.2f}%",
                                "Persentase Biaya Iklan terhadap Penjualan dari Iklan Langsung (ACOS Langsung)": f"{acos_direct:.2f}%",
                                "Jumlah Produk Dilihat": t_imp,
                                "Jumlah Klik Produk": t_cli,
                                "Persentase Klik Produk": f"{round(ctr, 2)}%"
                            }
                            
                            final_results.append(row_data)
                            
                        except Exception as e:
                            # Skip campaign yang error tapi lanjutkan yang lain
                            continue
                
                except Exception as e:
                    st.error(f"❌ Error batch {batch_num}: {str(e)}")
                    continue

            prog_bar.empty()
            status_text.empty()

            if not final_results:
                st.error("❌ Tidak ada data iklan yang berhasil diproses.")
                st.info("💡 Tips: Coba perpanjang range tanggal atau cek apakah ada campaign aktif.")
            else:
                # Buat DataFrame dengan kolom terurut
                df_ads = pd.DataFrame(final_results)
                
                # 🔴 PERBAIKAN: Urutkan kolom sesuai permintaan
                desired_columns = [
                    "Urutan", "Nama Iklan", "Status", "Jenis Iklan", "Kode Produk",
                    "Tampilan Iklan", "Mode Bidding", "Penempatan Iklan", "Tanggal Mulai",
                    "Tanggal Selesai", "Dilihat", "Jumlah Klik", "Persentase Klik",
                    "Konversi", "Konversi Langsung", "Tingkat konversi", "Tingkat Konversi Langsung",
                    "Biaya per Konversi", "Biaya per Konversi Langsung", "Produk Terjual",
                    "Terjual Langsung", "Omzet Penjualan", "Penjualan Langsung (GMV Langsung)",
                    "Biaya", "Efektifitas Iklan", "Efektivitas Langsung",
                    "Persentase Biaya Iklan terhadap Penjualan dari Iklan (ACOS)",
                    "Persentase Biaya Iklan terhadap Penjualan dari Iklan Langsung (ACOS Langsung)",
                    "Jumlah Produk Dilihat", "Jumlah Klik Produk", "Persentase Klik Produk"
                ]
                
                # Filter hanya kolom yang ada
                available_cols = [c for c in desired_columns if c in df_ads.columns]
                df_ads = df_ads[available_cols]
                
                # Tampilkan statistik
                total_biaya = df_ads["Biaya"].sum()
                total_omzet = df_ads["Omzet Penjualan"].sum()
                total_impression = df_ads["Dilihat"].sum()
                total_klik = df_ads["Jumlah Klik"].sum()
                
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total Campaign", len(df_ads))
                col2.metric("Total Biaya", f"Rp {total_biaya:,.0f}")
                col3.metric("Total Omzet", f"Rp {total_omzet:,.0f}")
                col4.metric("Total Klik", f"{total_klik:,.0f}")
                
                st.success(f"✅ Berhasil menarik {len(df_ads)} campaign dari {len(campaign_entries)} total campaign.")
                
                # Preview data
                st.subheader("📋 Preview Data Iklan")
                st.dataframe(df_ads, use_container_width=True, height=400)
                
                # Export Excel dengan formatting
                output_ads = io.BytesIO()
                with pd.ExcelWriter(output_ads, engine="openpyxl") as writer:
                    df_ads.to_excel(writer, index=False, sheet_name="Data Iklan")
                    
                    # Auto-adjust columns
                    worksheet = writer.sheets["Data Iklan"]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                excel_bytes = output_ads.getvalue()
                
                # Simpan ke database
                range_str = f"{start_ads} s/d {end_ads}"
                save_report_to_db(selected_shop_ads, f"ADS {range_str}", excel_bytes)
                
                col_dl, col_info = st.columns([1, 3])
                with col_dl:
                    st.download_button(
                        "⬇️ Download Excel", 
                        excel_bytes, 
                        f"Ads_Report_{selected_shop_ads}_{start_ads}_{end_ads}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col_info:
                    st.caption(f"💾 Data disimpan: ADS {range_str}")

        # ===============================
        # RIWAYAT IKLAN
        # ===============================
        st.divider()
        st.subheader("📜 Riwayat Laporan Iklan (Database)")

        history_ads = get_report_history(selected_shop_ads)

        if not history_ads:
            st.write("Belum ada riwayat laporan iklan.")
        else:
            # Filter hanya yang bertipe ADS
            ads_history = [item for item in history_ads if str(item["date_range"]).startswith("ADS")]
            
            if not ads_history:
                st.write("Belum ada riwayat laporan iklan.")
            else:
                for item in ads_history:
                    col1, col2, col3 = st.columns([3, 3, 2])
                    date_range_clean = item["date_range"].replace("ADS ", "")
                    col1.write(f"📅 {date_range_clean}")
                    col2.write(f"⏰ {item['created_at'][:19]}")
                    col3.download_button(
                        label="💾 Download",
                        data=item["csv_content"],
                        file_name=f"Ads_{selected_shop_ads}_{item['created_at'][:10]}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"ads_dl_{item['id']}"
                    )

with tab6:
    st.header("🔁 Seller Conversion")
    st.info("Untuk Menarik data seller conversion coba dengan Link ")

    
with tab7:
    st.header("🕒 Performa Iklan Per Jam")
    st.info("Data performa iklan seluruh toko berdasarkan jam (00:00 - 23:00) dengan timezone WIB (UTC+7).")

    now_jkt = datetime.now(WIB)
    yesterday_jkt = (now_jkt - timedelta(days=1)).date()

    if not shop_names:
        st.warning("Belum ada toko.")
    else:
        selected_shop_hourly = st.selectbox("Pilih Toko", shop_names, key="shop_hourly_v2")
        
        # ✅ PERBAIKAN: Pilihan preset tanggal
        st.subheader("📅 Pilih Tanggal")
        col_preset, col_custom = st.columns([1, 2])
        
        with col_preset:
            preset_hourly = st.selectbox(
                "Preset",
                ["Hari Ini", "Kemarin", "2 Hari Lalu", "Custom"],
                key="preset_hourly"
            )
        
        # ✅ Set default date berdasarkan preset menggunakan WIB
        today_wib = now_jkt.date()
        
        if preset_hourly == "Hari Ini":
            default_date = today_wib
        elif preset_hourly == "Kemarin":
            default_date = today_wib - dt.timedelta(days=1)
        elif preset_hourly == "2 Hari Lalu":
            default_date = today_wib - dt.timedelta(days=2)
        else:  # Custom
            default_date = today_wib - dt.timedelta(days=1)
        
        with col_custom:
            target_date = st.date_input(
                "Tanggal Target", 
                default_date, 
                key="date_hourly_v2",
                max_value=today_wib  # Tidak bisa pilih tanggal masa depan
            )
        
        # 🔴 PERBAIKAN: Info timezone WIB
        st.caption(f"🕐 Data akan diambil untuk: {target_date} (WIB - UTC+7)")
        
        # Validasi: tidak boleh tanggal masa depan
        if target_date > today_wib:
            st.error("❌ Tidak bisa mengambil data untuk tanggal masa depan!")
            st.stop()

        if st.button("🚀 Tarik Data Per Jam", type="primary"):
            token_row = get_shop_token(selected_shop_hourly)
            ACTIVE_SHOP_ID = token_row["shop_id"]
            ACTIVE_ACCESS_TOKEN = token_row["access_token"]

            # Progress indicator
            with st.spinner("🔄 Mengambil data performa per jam..."):
                # Format tanggal DD-MM-YYYY untuk API (tetap sama, API internal Shopee pakai UTC)
                date_str = target_date.strftime("%d-%m-%Y")
                ts_ads = int(time.time())

                path_hourly = "/api/v2/ads/get_all_cpc_ads_hourly_performance"
                sign_hourly = generate_sign_full(path_hourly, ts_ads, ACTIVE_ACCESS_TOKEN, ACTIVE_SHOP_ID)

                params_hourly = {
                    "partner_id": int(PARTNER_ID),
                    "timestamp": ts_ads,
                    "access_token": ACTIVE_ACCESS_TOKEN,
                    "shop_id": int(ACTIVE_SHOP_ID),
                    "sign": sign_hourly,
                    "performance_date": date_str  
                }

                try:
                    res_hourly = requests.get(
                        BASE_URL + path_hourly, 
                        params=params_hourly, 
                        timeout=30
                    ).json()
                except Exception as e:
                    st.error(f"❌ Error fetching data: {str(e)}")
                    st.stop()

                if res_hourly.get("error"):
                    error_msg = res_hourly.get("message", "Unknown error")
                    st.error(f"❌ Error dari Shopee API: {error_msg}")
                    
                    # Debug info
                    with st.expander("🔍 Debug Response"):
                        st.json(res_hourly)
                    st.stop()

                # Ambil data response
                hourly_list = res_hourly.get("response", [])

                if not hourly_list:
                    st.warning(f"⚠️ Tidak ada data iklan untuk tanggal {date_str} (WIB).")
                    st.info("💡 Tips: Coba pilih tanggal lain atau cek apakah ada iklan aktif pada tanggal tersebut.")
                    st.stop()
                
                # 🔴 PERBAIKAN: Proses data dengan timezone WIB
                # Buat template 24 jam (00:00 - 23:00 WIB)
                hourly_data_map = {}
                for h in range(24):
                    wib_time = f"{str(h).zfill(2)}:00"
                    hourly_data_map[wib_time] = {
                        "Jam WIB": wib_time,
                        "Lihat": 0,
                        "Klik": 0,
                        "Biaya": 0.0,
                        "CTR (%)": 0.0,
                        "CPC": 0.0
                    }

                total_impression = total_clicks = total_expense = 0
                
                for data in hourly_list:
                    h_num = data.get("hour")
                    if h_num is not None and 0 <= h_num <= 23:
                        # Konversi jam UTC ke WIB (UTC+7)
                        # Shopee API kembalikan jam dalam UTC, perlu konversi ke WIB
                        utc_hour = h_num
                        wib_hour = (utc_hour + 7) % 24  # Tambah 7 jam untuk WIB
                        # Handle kasus overflow (misal: UTC 20:00 -> WIB 03:00 pagi hari berikutnya)
                        if utc_hour + 7 >= 24:
                            # Data ini sebenarnya untuk hari berikutnya di WIB
                            # Tapi karena kita query per hari, biarkan saja atau skip
                            pass
                        
                        key = f"{str(wib_hour).zfill(2)}:00"
                        
                        impression = data.get("impression", 0) or 0
                        clicks = data.get("clicks", 0) or 0
                        expense = data.get("expense", 0) or 0
                        
                        hourly_data_map[key]["Lihat"] = impression
                        hourly_data_map[key]["Klik"] = clicks
                        hourly_data_map[key]["Biaya"] = expense
                        
                        # Hitung metrik turunan
                        ctr = (clicks / impression * 100) if impression > 0 else 0
                        cpc = (expense / clicks) if clicks > 0 else 0
                        
                        hourly_data_map[key]["CTR (%)"] = round(ctr, 2)
                        hourly_data_map[key]["CPC"] = round(cpc, 0)
                        
                        total_impression += impression
                        total_clicks += clicks
                        total_expense += expense

                # Susun DataFrame (urut berdasarkan jam 00-23)
                rows_hourly = []
                for h in range(24):
                    jam_key = f"{str(h).zfill(2)}:00"
                    rows_hourly.append(hourly_data_map[jam_key])

                df_hourly = pd.DataFrame(rows_hourly)
                
                # 🔴 PERBAIKAN: Hitung metrik tambahan
                total_ctr = (total_clicks / total_impression * 100) if total_impression > 0 else 0
                total_cpc = (total_expense / total_clicks) if total_clicks > 0 else 0
                
                # Cari peak hour
                peak_hour_idx = df_hourly["Biaya"].idxmax()
                peak_hour = df_hourly.loc[peak_hour_idx, "Jam WIB"]
                peak_biaya = df_hourly.loc[peak_hour_idx, "Biaya"]
                
                peak_click_idx = df_hourly["Klik"].idxmax()
                peak_click_hour = df_hourly.loc[peak_click_idx, "Jam WIB"]
                peak_click_val = df_hourly.loc[peak_click_idx, "Klik"]

            # Tampilkan ringkasan
            st.success(f"✅ Berhasil mengambil data performa per jam untuk {target_date} (WIB)")
            
            # 🔴 PERBAIKAN: Metrics cards
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Lihat", f"{total_impression:,}")
            col2.metric("Total Klik", f"{total_clicks:,}")
            col3.metric("Total Biaya", f"Rp {total_expense:,.0f}")
            col4.metric("Total CTR", f"{total_ctr:.2f}%")
            
            col5, col6 = st.columns(2)
            col5.metric("Peak Hour (Biaya)", f"{peak_hour} ({peak_biaya:,.0f})")
            col6.metric("Peak Hour (Klik)", f"{peak_click_hour} ({peak_click_val:,})")

            # 🔴 PERBAIKAN: Visualisasi Chart
            st.subheader("📊 Grafik Performa Per Jam")
            
            tab_chart1, tab_chart2, tab_chart3 = st.tabs(["Biaya", "Klik & Lihat", "CTR"])
            
            with tab_chart1:
                st.bar_chart(df_hourly.set_index("Jam WIB")["Biaya"])
            
            with tab_chart2:
                chart_data = df_hourly.set_index("Jam WIB")[["Klik", "Lihat"]]
                st.line_chart(chart_data)
            
            with tab_chart3:
                st.line_chart(df_hourly.set_index("Jam WIB")["CTR (%)"])

            # Tampilkan tabel data
            st.subheader(f"📋 Detail Data Per Jam ({target_date} WIB)")
            
            # 🔴 PERBAIKAN: Highlight row dengan aktivitas tinggi
            def highlight_peak(row):
                if row["Biaya"] == peak_biaya:
                    return ['background-color: #ffeb3b'] * len(row)
                elif row["Biaya"] > (total_expense / 24 * 1.5):  # Di atas rata-rata 50%
                    return ['background-color: #e3f2fd'] * len(row)
                return [''] * len(row)
            
            styled_df = df_hourly.style.apply(highlight_peak, axis=1).format({
                "Biaya": "Rp {:,.0f}",
                "CPC": "Rp {:,.0f}",
                "CTR (%)": "{:.2f}%",
                "Lihat": "{:,}",
                "Klik": "{:,}"
            })
            
            st.dataframe(styled_df, use_container_width=True, height=600)

            # 🔴 PERBAIKAN: Export Excel dengan formatting
            output_h = io.BytesIO()
            with pd.ExcelWriter(output_h, engine="openpyxl") as writer:
                # Sheet 1: Data Per Jam
                df_hourly.to_excel(writer, index=False, sheet_name="Hourly_Performance")
                
                # Sheet 2: Summary
                summary_data = {
                    "Metrik": [
                        "Tanggal (WIB)",
                        "Total Jam dengan Aktivitas",
                        "Total Lihat (Impression)",
                        "Total Klik",
                        "Total Biaya",
                        "Rata-rata CTR",
                        "Rata-rata CPC",
                        "Peak Hour (Biaya Tertinggi)",
                        "Peak Hour (Klik Tertinggi)",
                        "Jam Tanpa Aktivitas"
                    ],
                    "Nilai": [
                        str(target_date),
                        len([h for h in hourly_list if h.get("impression", 0) > 0]),
                        total_impression,
                        total_clicks,
                        total_expense,
                        f"{total_ctr:.2f}%",
                        f"Rp {total_cpc:,.0f}",
                        f"{peak_hour} (Rp {peak_biaya:,.0f})",
                        f"{peak_click_hour} ({peak_click_val:,} klik)",
                        len([h for h in hourly_list if h.get("impression", 0) == 0])
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name="Summary")
                
                # Formatting
                for sheet_name in ["Hourly_Performance", "Summary"]:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_bytes = output_h.getvalue()
            
            # Simpan ke database
            save_report_to_db(
                selected_shop_hourly, 
                f"HOURLY {target_date}", 
                excel_bytes
            )
            
            col_dl, col_info = st.columns([1, 3])
            with col_dl:
                st.download_button(
                    label="💾 Download Excel",
                    data=excel_bytes,
                    file_name=f"Ads_Hourly_{selected_shop_hourly}_{target_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_info:
                st.caption(f"💾 Data disimpan: HOURLY {target_date} (WIB)")

        # ===============================
        # RIWAYAT PER JAM
        # ===============================
        st.divider()
        st.subheader("📜 Riwayat Laporan Per Jam (Database)")

        if 'selected_shop_hourly' in locals():
            history_hourly = get_report_history(selected_shop_hourly)
            
            if not history_hourly:
                st.write("Belum ada riwayat laporan per jam.")
            else:
                # Filter hanya yang bertipe HOURLY
                hourly_history = [item for item in history_hourly if str(item["date_range"]).startswith("HOURLY")]
                
                if not hourly_history:
                    st.write("Belum ada riwayat laporan per jam.")
                else:
                    for item in hourly_history:
                        col1, col2, col3 = st.columns([3, 3, 2])
                        date_clean = item["date_range"].replace("HOURLY ", "")
                        col1.write(f"📅 {date_clean} (WIB)")
                        col2.write(f"⏰ {item['created_at'][:19]}")
                        col3.download_button(
                            label="💾 Download",
                            data=item["csv_content"],
                            file_name=f"Hourly_{selected_shop_hourly}_{date_clean}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"hourly_dl_{item['id']}"
                        )

